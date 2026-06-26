"""
export_db_to_excel.py  —  Export the ChromaDB knowledge base to a formatted Excel file.

Output: knowledge_base_export.xlsx  (saved to project root)

Sheet 1  "Knowledge Base"  — all 403 items, colour-coded by tag
Sheet 2  "Summary"         — counts by tag and source type
"""

import chromadb
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Config ─────────────────────────────────────────────────────────────────────
DB_PATH         = "./db"
COLLECTION_NAME = "logo_plc"
OUTPUT_FILE     = "knowledge_base_export.xlsx"
TEXT_PREVIEW    = 300   # characters of text to show in the table

# ── Colour palette (hex fill colours, no leading #) ───────────────────────────
COLOURS = {
    "SPEC": "FFD966",   # amber
    "GATE": "4BACC6",   # teal
    "LD":   "8064A2",   # purple  (text will be white)
    "ST":   "70AD47",   # green
}

# Tags where white text is more readable on the dark background
DARK_TAGS = {"LD"}

HEADER_FILL  = PatternFill("solid", fgColor="2E4057")   # dark navy
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(size=10)
BODY_FONT_WH = Font(size=10, color="FFFFFF")            # white text on dark rows

THIN = Side(border_style="thin", color="D0D0D0")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical="top")
TOP  = Alignment(vertical="top")


# ── Helpers ────────────────────────────────────────────────────────────────────

def set_header(ws, headers, col_widths):
    """Write a styled header row."""
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = TOP
        ws.column_dimensions[get_column_letter(col)].width = width


def style_cell(cell, tag, wrap=False):
    """Apply tag colour fill and appropriate text colour to a data cell."""
    hex_colour = COLOURS.get(tag, "FFFFFF")
    cell.fill   = PatternFill("solid", fgColor=hex_colour)
    cell.font   = BODY_FONT_WH if tag in DARK_TAGS else BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = WRAP if wrap else TOP


# ── Main export ────────────────────────────────────────────────────────────────

def export():
    print("Connecting to ChromaDB...")
    client     = chromadb.PersistentClient(path=DB_PATH)
    collection = client.get_collection(COLLECTION_NAME)

    print(f"Fetching all {collection.count()} items...")
    # Fetch everything in one call — no query needed, just get all
    all_items = collection.get(include=["documents", "metadatas"])

    ids   = all_items["ids"]
    docs  = all_items["documents"]
    metas = all_items["metadatas"]

    # Sort by type then tag then id for a logical reading order
    combined = sorted(
        zip(ids, docs, metas),
        key=lambda x: (x[2].get("type", ""), x[2].get("tag", ""), x[0])
    )

    wb = Workbook()

    # ── Sheet 1: Knowledge Base ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Knowledge Base"

    headers    = ["#", "Source", "Tag", "Type", "Text Content (first 300 chars)"]
    col_widths = [5,   28,       8,     18,      90]
    set_header(ws1, headers, col_widths)
    ws1.row_dimensions[1].height = 22

    # Counters for Summary sheet
    tag_counts  = {"SPEC": 0, "GATE": 0, "LD": 0, "ST": 0}
    type_counts = {"manual": 0, "annotation": 0, "spec_annotation": 0}

    for row_num, (item_id, doc, meta) in enumerate(combined, start=2):
        tag       = meta.get("tag",    "")
        item_type = meta.get("type",   "")
        source    = meta.get("source", item_id)
        preview   = doc[:TEXT_PREVIEW].replace("\n", " ")
        if len(doc) > TEXT_PREVIEW:
            preview += "..."

        values = [row_num - 1, source, tag, item_type, preview]
        for col, val in enumerate(values, start=1):
            cell = ws1.cell(row=row_num, column=col, value=val)
            style_cell(cell, tag, wrap=(col == 5))

        ws1.row_dimensions[row_num].height = 60

        if tag in tag_counts:
            tag_counts[tag] += 1
        if item_type in type_counts:
            type_counts[item_type] += 1

    # Freeze the header row so it stays visible while scrolling
    ws1.freeze_panes = "A2"

    # ── Sheet 2: Summary ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Summary")

    # --- Section 1: by tag ---
    ws2["A1"] = "Counts by Tag"
    ws2["A1"].font = Font(bold=True, size=13)

    tag_headers = ["Tag", "Colour", "Item Count", "Description"]
    tag_widths  = [10,    12,       14,            48]
    set_header_at(ws2, row=2, headers=tag_headers, col_widths=tag_widths)

    tag_descriptions = {
        "SPEC": "Hardware specs, wiring, technical data, power supply, temperature",
        "GATE": "Logic gates, special functions (timers, counters, comparator)",
        "LD":   "Ladder/circuit diagram content, programming rules",
        "ST":   "Timing diagrams, status tables, operating states",
    }
    for i, (tag, count) in enumerate(sorted(tag_counts.items()), start=3):
        ws2.cell(row=i, column=1, value=tag)
        colour_cell = ws2.cell(row=i, column=2, value=tag)
        colour_cell.fill = PatternFill("solid", fgColor=COLOURS.get(tag, "FFFFFF"))
        colour_cell.font = BODY_FONT_WH if tag in DARK_TAGS else BODY_FONT
        ws2.cell(row=i, column=3, value=count)
        ws2.cell(row=i, column=4, value=tag_descriptions.get(tag, ""))
        for col in range(1, 5):
            ws2.cell(row=i, column=col).border = THIN_BORDER

    total_row = 3 + len(tag_counts)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=3, value=sum(tag_counts.values())).font = Font(bold=True)

    # --- Section 2: by source type ---
    start_row = total_row + 3
    ws2.cell(row=start_row, column=1).value = "Counts by Source Type"
    ws2.cell(row=start_row, column=1).font  = Font(bold=True, size=13)

    type_headers = ["Source Type", "Item Count", "Description"]
    type_widths  = [20,            14,            55]
    set_header_at(ws2, row=start_row + 1, headers=type_headers, col_widths=type_widths)

    type_descriptions = {
        "manual":          "Chunks extracted from the LOGO! 8 system manual PDF (388 pages)",
        "annotation":      "Hand-crafted ladder circuit examples covering all 18 LOGO!JSON block types",
        "spec_annotation": "Hand-crafted parameter table entries (timers, counters, voltage, temperature)",
    }
    for i, (t, count) in enumerate(sorted(type_counts.items()), start=start_row + 2):
        ws2.cell(row=i, column=1, value=t)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=type_descriptions.get(t, ""))
        for col in range(1, 4):
            ws2.cell(row=i, column=col).border = THIN_BORDER

    total_row2 = start_row + 2 + len(type_counts)
    ws2.cell(row=total_row2, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row2, column=2, value=sum(type_counts.values())).font = Font(bold=True)

    # --- Section 3: note ---
    note_row = total_row2 + 3
    ws2.cell(row=note_row, column=1).value = "Notes"
    ws2.cell(row=note_row, column=1).font  = Font(bold=True, size=13)
    ws2.cell(row=note_row + 1, column=1).value = (
        "Some manual pages appear more than once because they carry multiple tags "
        "(e.g. page 136 is tagged both GATE and LD). This is intentional — it allows "
        "the same page to be retrieved by both gate queries and ladder diagram queries."
    )
    ws2.cell(row=note_row + 1, column=1).alignment = Alignment(wrap_text=True)
    ws2.merge_cells(
        start_row=note_row + 1, start_column=1,
        end_row=note_row + 3,   end_column=4
    )
    ws2.row_dimensions[note_row + 1].height = 60

    # ── Save ───────────────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print()
    print(f"Saved: {OUTPUT_FILE}")
    print()
    print("Sheet 1  'Knowledge Base':")
    print(f"  {len(combined)} rows, colour-coded by tag")
    print(f"  SPEC (amber):   {tag_counts['SPEC']}")
    print(f"  GATE (teal):    {tag_counts['GATE']}")
    print(f"  LD   (purple):  {tag_counts['LD']}")
    print(f"  ST   (green):   {tag_counts['ST']}")
    print()
    print("Sheet 2  'Summary':")
    print(f"  manual:          {type_counts['manual']}")
    print(f"  annotation:      {type_counts['annotation']}")
    print(f"  spec_annotation: {type_counts['spec_annotation']}")
    print(f"  TOTAL:           {sum(type_counts.values())}")


def set_header_at(ws, row, headers, col_widths):
    """Write a styled header row at a specific row number."""
    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.border    = THIN_BORDER
        cell.alignment = TOP
        ws.column_dimensions[get_column_letter(col)].width = width


if __name__ == "__main__":
    export()
